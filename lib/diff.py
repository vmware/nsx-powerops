#!/usr/bin/env python
# coding: utf-8
#############################################################################################################################################################################################
#                                                                                                                                                                                           #
# NSX-T Power Operations                                                                                                                                                                    #
#                                                                                                                                                                                           #
# Copyright 2020 VMware, Inc.  All rights reserved				                                                                                                                            #
#                                                                                                                                                                                           #
# The MIT license (the “License”) set forth below applies to all parts of the NSX Power Operations project.  You may not use this file except in compliance with the License.               #
#                                                                                                                                                                                           #
# MIT License                                                                                                                                                                               #
#                                                                                                                                                                                           #
# Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"),                                        #
# to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense,                                        #
# and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:                                                #
#                                                                                                                                                                                           #
# The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.                                                            #
#                                                                                                                                                                                           #
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,                                       #
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,                             #
# WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.                                #
#                                                                                                                                                                                           #
# *--------------------------------------------------------------------------------------* #                                                                                                #
# **************************************************************************************** #                                                                                                #
#   VMware NSX-T PowerOps by @dominicfoley & @stephensauer                                 #                                                                                                #
#   A day 2 operations tool for helping to document and healthcheck an NSX-T environment   #                                                                                                #
# **************************************************************************************** #                                                                                                #
# *--------------------------------------------------------------------------------------* #                                                                                                #
#                                                                                                                                                                                           #
#############################################################################################################################################################################################
import sys, os, datetime, time
from lib.excel import CreateXLSFile
from lib.docs_summary import SheetSummary
from lib.docs_alarms import SheetAlarms
from lib.docs_groups import SheetSecGrp
from lib.docs_securitypolicies import SheetSecPol
from lib.docs_securitypolicies_and_rules import SheetSecDFW
from lib.docs_tier1_segments import SheetT1Segments
from lib.docs_lr_summary import SheetRouterSum
from lib.docs_lr_ports import SheetRouterPorts
from lib.docs_tier1_segments import SheetT1Segments
from lib.docs_logical_switches import SheetSegments
from lib.docs_RoutingSessions import SheetBGPSession
from lib.docs_tier0_routingtables import SheetT0RoutingTable
from lib.docs_tier1_forwardingtables import SheetT1ForwardingTable
from lib.docs_nsxmanagers import SheetNSXManagerInfo
from lib.docs_discovered_nodes import SheetFabDiscoveredNodes
from lib.docs_transportzones import SheetTZ
from lib.docs_services import SheetNSXServices
from lib.docs_tn_tunnels import SheetTunnels
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


DIFF = None
# Params for diff function
def SetDiffFileName(val):
    global DIFF
    DIFF = val

def GetDiffFileName():
    global DIFF
    return DIFF
    
def IfDiff():
    global DIFF
    if DIFF == None:
        return False
    else:
        return True
    
def SetXLSDiffFile(auth_list, xls_diff_bkp_filename):
    # Workbook for current config
    global NSX_Config
    NSX_Config = {}
    start_time = time.time()
    WORKBOOK = CreateXLSFile(auth_list,'Audit_DIFF_NSX')
    # Workbook object for reference diff bkp file
    DIFF_WORKBOOK = load_workbook(xls_diff_bkp_filename)
    if WORKBOOK != None:
        # Generating Summary
        print('\nGenerating Summary sheet')
        TN_WS = WORKBOOK[0].active
        TN_WS.title = "Summary"
        SheetSummary(auth_list,WORKBOOK[0],TN_WS,NSX_Config)
        # Generating Information sheet
        print('Generating NSX-T Manager Information sheet')
        TN_WS = WORKBOOK[0].create_sheet("NSX_Manager_Info")
        SheetNSXManagerInfo(auth_list,WORKBOOK[0],TN_WS,NSX_Config)
        # Generating NSX-T Fabric Discovered Nodes sheet
        print('Generating NSX-T Fabric Discovered Nodes sheet')
        TN_WS = WORKBOOK[0].create_sheet("Transport_Nodes")
        SheetFabDiscoveredNodes(auth_list,WORKBOOK[0],TN_WS,NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Transport_Nodes'])
        # Generating NSX-T Transport Zones sheet
        print('Generating NSX-T Transport Zones sheet')
        TN_WS = WORKBOOK[0].create_sheet("Transport_Zones")
        SheetTZ(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Transport_Zones'])
        # Generating NSX-T Services sheet
        print('Generating NSX-T Services sheet')
        TN_WS = WORKBOOK[0].create_sheet("Services")
        SheetNSXServices(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Services'])
        # Generating NSX-T TEP Tunnels sheet
        print('Generating NSX-T TEP Tunnels sheet')
        TN_WS = WORKBOOK[0].create_sheet("Transport_Node_Tunnels")
        SheetTunnels(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Transport_Node_Tunnels'])
        # Generating NSX-T Segments sheet
        print('Generating NSX-T Segments sheet')
        TN_WS = WORKBOOK[0].create_sheet("Segments")
        SheetSegments(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Segments'])
        # Generating NSX-T Router Summary sheet
        print('Generating NSX-T Router Summary sheet')
        TN_WS = WORKBOOK[0].create_sheet("Logical_Router_Summary")
        SheetRouterSum(auth_list,WORKBOOK[0],TN_WS)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Logical_Router_Summary'])
        # Generating NSX-T Router Ports sheet
        print('Generating NSX-T Router Ports sheet')
        TN_WS = WORKBOOK[0].create_sheet("Logical_Router_Ports")
        SheetRouterPorts(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Logical_Router_Ports'])
        # Generating NSX-T Router T1 Segments sheet
        print('Generating NSX-T Router T1 Segments sheet')
        TN_WS = WORKBOOK[0].create_sheet("Tier1_Segments")
        SheetT1Segments(auth_list,WORKBOOK[0],TN_WS,NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Tier1_Segments'])
        # Generate XLS Tab for Router
        print('Generating NSX-T Router T0 Routing Tables sheet')
        TN_WS = WORKBOOK[0].create_sheet("Tier0_Routing_Tables")
        SheetT0RoutingTable(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Tier0_Routing_Tables'])
        # Generating NSX-T Router T0 BGP Sessions sheet
        print('Generating NSX-T Router T0 BGP Sessions sheet')
        TN_WS = WORKBOOK[0].create_sheet("Tier0_BGP_Sessions")
        SheetBGPSession(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Tier0_BGP_Sessions'])
        # Generating NSX-T Router T1 Forwarding Tables sheet
        print('Generating NSX-T Router T1 Forwarding Tables sheet')
        TN_WS = WORKBOOK[0].create_sheet("Tier1_Forwarding_Tables")
        SheetT1ForwardingTable(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Tier1_Forwarding_Tables'])
        # Generating NSX-T Groups sheet
        print('Generating NSX-T Groups sheet')
        TN_WS = WORKBOOK[0].create_sheet("Security_Groups")
        SheetSecGrp(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Security_Groups'])
        # Generating NSX-T Security Policies sheet
        print('Generating NSX-T Security Policies sheet')
        TN_WS = WORKBOOK[0].create_sheet("Security_Policies")
        SheetSecPol(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Security_Policies'])
        # Generating NSX-T Security DFW sheet
        print('Generating NSX-T Security DFW sheet')
        TN_WS = WORKBOOK[0].create_sheet("Rules_Distributed_Firewall")
        SheetSecDFW(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
        CheckXLSTabDiff(TN_WS, DIFF_WORKBOOK['Rules_Distributed_Firewall'])
        # Generating NSX-T Alarms sheet
        print('Generating NSX-T Alarms sheet')
        TN_WS = WORKBOOK[0].create_sheet("Alarms")
        SheetAlarms(auth_list,WORKBOOK[0],TN_WS, NSX_Config)
       
        # Save XLS file
        print("\nDocumentation \"Audit_DIFF\" set took %s seconds to complete\n" % (time.time() - start_time))
        WORKBOOK[0].save(WORKBOOK[1])

def CheckXLSTabDiff(WS_CURRENT, WS_REFERENCE):
    # For each line in WS_CURRENT
    ## => Check if exist in WS_REFERENCE.
    ##     => If NO THEN color line in YELLOW/ORANGE (because new/edited)
    ##     => If YES : no modification and remove the line in WS_REFERENCE
    ## => At the end, if WS_REFERENCE is not empty : (deleted objects)
    ##      => Then add these lines in WS_CURRENT and fill line in RED (because deleted)
    start_data_line = 2
    cur_data_line = 2
    end_data_line = WS_CURRENT.max_row
    row_len = WS_CURRENT.max_column
    list_unchanged_rows = []
    for row in WS_CURRENT.iter_rows(min_row=start_data_line, min_col=1, max_row=end_data_line, max_col=row_len, values_only=True):
        # Convert tuple to list for comparison
        row = list(row)
        found = False
        i = 2
        ws_ref_num_line = WS_REFERENCE.max_row
        ws_ref_row_len = WS_REFERENCE.max_column
        # Replace '' by None if empty
        for x in range(0,len(row)):
            if row[x] == '':
                row[x] = None
        for ref_row in WS_REFERENCE.iter_rows(min_row=start_data_line, min_col=1, max_row=ws_ref_num_line, max_col=ws_ref_row_len, values_only=True):
            # Convert tuple to list for comparison
            ref_row = list(ref_row)
            # print ("CUR_ROW = "+str(row))
            # print ("REF_ROW = "+str(ref_row))
            if row == ref_row:
                found = True
                WS_REFERENCE.delete_rows(i)
                #print (ref_row)
                #print ("Deleted REF" + str(list(WS_REFERENCE.iter_rows(min_row=i, min_col=1, max_row=i, max_col=ws_ref_row_len))))
                break
            i = i + 1
        # Style : Fill cell with orange fil
        if not found:
            for row_obj in WS_CURRENT.iter_rows(min_row=cur_data_line, min_col=1, max_row=cur_data_line, max_col=row_len):
                for cell_obj in row_obj:
                    WS_CURRENT[cell_obj.coordinate].fill = PatternFill(start_color='00FFCC00', end_color='00FFCC00', fill_type='solid')
                    WS_CURRENT.cell(row=cur_data_line, column=row_len+1).value = "MODIFIED"
                    WS_CURRENT.cell(row=cur_data_line, column=row_len+1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    WS_CURRENT.cell(row=cur_data_line, column=row_len+1).font = Font (italic=True)
                    
        else:
            list_unchanged_rows.append(cur_data_line)
        cur_data_line = cur_data_line + 1
    
    ## Remove unchanged rows in diff output file
    for val in list_unchanged_rows:
        WS_CURRENT.cell(row=val, column=row_len+1).value = "UNCHANGED"
        WS_CURRENT.cell(row=val, column=row_len+1).alignment = Alignment(horizontal='left', vertical='center')
        WS_CURRENT.cell(row=val, column=row_len+1).font = Font (italic=True)

    ### At the end loop in ref_Workbook and add in red + stripped font these rows in cur_WB (because deleted)
    for ref_left_row in WS_REFERENCE.iter_rows(min_row=start_data_line, min_col=1, max_row=ws_ref_num_line, max_col=ws_ref_row_len, values_only=True):
        if not all(elem is None for elem in list(ref_left_row)):
            WS_CURRENT.append(ref_left_row)
            for x in range(1, ws_ref_row_len+1):
                WS_CURRENT.cell(row=WS_CURRENT.max_row, column=x).fill = PatternFill(start_color='00FF3333', end_color='00FF3333', fill_type='solid')
                WS_CURRENT.cell(row=WS_CURRENT.max_row, column=x).font = Font(italic=True, strike=True)
            WS_CURRENT.cell(row=WS_CURRENT.max_row, column=ws_ref_row_len+1).value = "DELETED"
            WS_CURRENT.cell(row=WS_CURRENT.max_row, column=ws_ref_row_len+1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            WS_CURRENT.cell(row=WS_CURRENT.max_row, column=ws_ref_row_len+1).font = Font (italic=True)
    
    # Add Title to last column (diff status : Unchanged, Modified, Deleted)
    WS_CURRENT.cell(row=1, column=row_len+1).value = "Diff Status"
    # Add last column Filter
    WS_CURRENT.auto_filter.ref = get_column_letter(row_len+1) + "1:" + get_column_letter(row_len+1) + str(WS_CURRENT.max_row)
    # filter last column to : Modified & Deleted
    WS_CURRENT.column_dimensions[get_column_letter(row_len+1)].auto_size = True
    # Hide the last column with diff status
    #WS_CURRENT.column_dimensions[get_column_letter(row_len+1)].hidden= True

   