#!/usr/bin/env python
# coding: utf-8
#
# Librairies Word/Excel/ CSV
import openpyxl, os, datetime, pathlib, re, csv
from lib.system import style, GetCSV
import lib.menu
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Font, Fill, Color, PatternFill, Border, Side, colors, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.utils import get_column_letter


def CreateXLSFile(auth_list,XLS_Files,SheetFunction = None):
	# Creation of CSV files
	if GetCSV():
		now = datetime.datetime.now()
		DateString = now.strftime("%Y%m%d")
		OUTPUT_CSV = lib.menu.XLS_Dest + os.sep + XLS_Files + "_"  + DateString + ".csv"
		fname = pathlib.Path(OUTPUT_CSV)
		if fname.exists():
			print(str(fname) + style.RED + '\n==> File already exists. Not attempting to overwite' + style.NORMAL + "\n")
			return None
		print('\nGenerating NSX-T Manager output: ' + style.ORANGE + XLS_Files + "_"  + DateString + ".csv" + style.NORMAL + '\n')
		with open(OUTPUT_CSV, 'w', newline="") as file:
			if SheetFunction != None:
				WORKBOOK = Workbook()
				TN_WS = WORKBOOK.active
				SheetFunction(auth_list,csv.writer(file),TN_WS)
			
			return [csv.writer(file), OUTPUT_CSV]
	else:	
		# Creation of Excel files
		now = datetime.datetime.now()
		DateString = now.strftime("%Y%m%d")
		OUTPUT_XLS = lib.menu.XLS_Dest + os.sep + XLS_Files + "_"  + DateString + ".xlsx"
		fname = pathlib.Path(OUTPUT_XLS)
		if fname.exists():
			print(str(fname) + style.RED + '\n==> File already exists. Not attempting to overwite' + style.NORMAL + "\n")
			return None
		
		print('\nGenerating NSX-T Manager output: ' + style.ORANGE + XLS_Files + "_"  + DateString + ".xlsx" + style.NORMAL + '\n')
		WORKBOOK = Workbook()
		if SheetFunction != None:
			TN_WS = WORKBOOK.active
			TN_WS.title = XLS_Files
			SheetFunction(auth_list,WORKBOOK,TN_WS)
			WORKBOOK.save(OUTPUT_XLS)

		return [WORKBOOK, OUTPUT_XLS]

# Create Sheet XLS
# ------------------------------------------------------------------------------------------
def FillSheet(TN_WB,title,Header_row,List_row,color, StyleTable = "TableStyleLight9", FirstColumn = False, start_cell = 'A1'):
	"""
	CreateSheet(TN_WB,title,Header_row,List_row,color)
	Create a Excel sheet
	Parameters
	----------
	TN_WB : object
	    workbook Excel object
	title : str
	    Title of tab (no spaces, no specials caracters)
	Header_row : tuple
	    Tuple with all colums name (to be in order)
	List_row : tuple
	    List of list with all informations of table
	color : str
	    Color of tab in Hexa (ex: "0FF0FF")
	"""
	# Get XLS sheet tab
	TN_WS = TN_WB.get_sheet_by_name(title)
	TN_WS.sheet_properties.tabColor = color
	TN_WS.title = title
	# Create of header
	TN_WS.append(Header_row)
	for line in List_row:
		TN_WS.append(line)

	ColumnLetter = ""
	# Treatment of Start line of sheet
	start_line = str(len(List_row ) + 1)
	if start_cell != "A1":
		start_line = str(int(re.findall(r'\d+', start_cell)[0])+ len(List_row ))

	n = len(Header_row)
	while n > 0:
		n, remainder = divmod(n - 1, 26)
		ColumnLetter = chr(65 + remainder) + ColumnLetter
	Range = start_cell + ":" + ColumnLetter + start_line
	FormatSheet(TN_WS, Range, StyleTable, FirstColumn)

# Create CSV sheet
# ------------------------------------------------------------------------------------------
def FillSheetCSV(CSV,Header_row,List_row):
	CSV.writerow(Header_row)
	for r in List_row:
		# if list (with \n in a cell) => convert \n to space
		new_row = []
		for c in r:
			c = str(c)
			new_row.append(c.replace("\n"," "))
		CSV.writerow(new_row)

# Conditionnal Formating 
# ------------------------------------------------------------------------------------------
def ConditionnalFormat(Onglet, Range, SearchText, Line=False, Color = None):
	DefColor ={}
	# Colors definitions
	DefColor['RED'] = {
		'pattern_fill' : PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type='solid'),
		'text': Font(color="9C0103"),
		'fill': PatternFill(bgColor="FFC7CE")
	}
	DefColor['GREEN'] = {
		'pattern_fill' : PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type='solid'),
		'text': Font(color="006100"),
		'fill': PatternFill(bgColor="C6EFCE")
	}
	DefColor['ORANGE'] = {
		'pattern_fill' : PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type='solid'),
		'text': Font(color="9C5700"),
		'fill': PatternFill(bgColor="FFEB9C")
	}
	# Conditionnal rules
	# Red if not find - green if found
	if Color is None and Line is False:
		Onglet.conditional_formatting.add(Range, FormulaRule(formula=[f'ISERROR(SEARCH("{SearchText}",{Range}))'], stopIfTrue=True, fill=DefColor['RED']['fill'], font=DefColor['RED']['text']))
		Onglet.conditional_formatting.add(Range, FormulaRule(formula=[f'NOT(ISERROR(SEARCH("{SearchText}",{Range})))'], stopIfTrue=True, fill=DefColor['GREEN']['fill'], font=DefColor['GREEN']['text']))
	elif Color is not None and Line is False:
		Onglet.conditional_formatting.add(Range, FormulaRule(formula=[f'NOT(ISERROR(SEARCH("{SearchText}",{Range})))'], stopIfTrue=True, fill=DefColor[Color]['fill'], font=DefColor[Color]['text']))
	elif Color is not None and Line == True:
		for row in range(2,Onglet.max_row + 1):
			if Onglet.cell(row=row, column=ord(Range.lower())- 96).value == SearchText:
				for row_cells in Onglet.iter_rows(min_row = row,  max_row = row):
					for cell in row_cells:
						cell.fill = DefColor[Color]['pattern_fill']
						cell.font = DefColor[Color]['text']


# Format XLS sheet
# ------------------------------------------------------------------------------------------
def FormatSheet(Onglet, Range, StyleTable = "TableStyleLight9", FirstColumn = False):
	"""
	FormatTableau(Onglet, Range)
	Format a sheet in style (for all styles see excel)
	Parameters
	----------
	Onglet : str
	    Tab to format
	Range : 
	    Range of sheet (ex: "A1:G23")
	"""
	for col in Onglet.columns:
		max_length = 0
		adjusted_width = 0
		for cell in col:
			# Necessary to avoid error on empty cells
			try:
				if "\n" in str(cell.value): max_length = len(max(str(cell.value).split('\n'), key=len))
				if len(str(cell.value)) > max_length: max_length = len(cell.value)
				cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
			except:
				pass	
		if max_length >= 40: max_length = 40
		adjusted_width = (max_length + 2) * 1.1
		Onglet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
	
	ST = TableStyleInfo(name=StyleTable, showFirstColumn=FirstColumn,showLastColumn=False, showRowStripes=False, showColumnStripes=True)
	tab = Table(displayName=Onglet.title, ref=Range, tableStyleInfo=ST)
	Onglet.add_table(tab)
	print("-- --> Writing tab " + style.ORANGE + Onglet.title + style.NORMAL)
